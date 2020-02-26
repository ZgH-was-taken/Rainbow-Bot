import discord
from discord.ext import commands
from discord.utils import get
import random as rnd
import openpyxl
import asyncio
import datetime as dt


with open('token.txt', 'r') as f:
    token = f.readline()

with open('ids.txt', 'r') as f:
    pronounMsg = int(f.readline()[:-1])
    sexualityMsg = int(f.readline()[:-1])
    genderMsg = int(f.readline()[:-1])
    enginMsg = int(f.readline())

client = discord.Client()
bot = commands.Bot(command_prefix='!', case_insensitive = True)
bot.remove_command('help')

wb = openpyxl.load_workbook('Role Colours.xlsx')
ws = wb['Sheet 1']


@bot.event
async def on_ready():
    global server, generalChannel, ruleChannel, botChannel, execChannel, botMember
    server = get(bot.guilds, name='Rainbow Engineering')
    generalChannel = get(server.channels, name='✨-general')
    ruleChannel = get(server.channels, name='📜-rules-and-info')
    botChannel = get(server.channels, name='🤖-bot-commands')
    execChannel = get(server.channels, name='exec-chat')
    botMember = get(server.members, id=bot.user.id)
    print('Bot is ready')




helpMessage = '''\n```
All commands are used in #🤖-bot-commands
!help      Displays this message

!roles      Displays a list of roles you can self-assign/remove

`[roleName]      Gives/takes away a role
      Some roles can also be given by reacting in #📜-rules-and-info

!suggest [name] [colour]      Suggests a new role to execs with a name and
      optional hex colour in the form rgb/rrggbb
      Role names with multiple words should be suggested using " "
```\n'''
@bot.command()
async def help(ctx):
    await botChannel.send(helpMessage)


@bot.command()
async def roles(ctx):
    if ctx.channel is not botChannel: return
    roleMessage = '\n```\n'
    upper = get(server.roles,name='Ask-Pronouns').position
    middle = get(server.roles,name='Questioning').position
    lower = get(server.roles,name='Software').position
    for i in range(len(server.roles)-1,0,-1):
        if server.roles[i].position <= upper and server.roles[i].position >= middle:
            roleMessage += server.roles[i].name + '  '
    roleMessage += '\n\n'
    for i in range(len(server.roles)-1,0,-1):
        if server.roles[i].position < middle and server.roles[i].position >= lower:
            roleMessage += server.roles[i].name + '  '
    roleMessage += '\n\n'
    i = 1
    while ws.cell(i,1).value is not None:
        roleMessage += ws.cell(i,1).value + '  '
        i += 1
    roleMessage += '\n```\n'
    await botChannel.send(roleMessage)


@bot.command()
async def suggest(ctx, name, hex='0x0'):
    if ctx.channel is not botChannel: return
    msg = await execChannel.send(ctx.author.name + ' has suggested a role \'' + name + '\', with hex colour ' + hex)
    await msg.add_reaction('✅')
    await msg.add_reaction('⛔')
    msg = await execChannel.fetch_message(msg.id)
    def check(reaction, member):
        return (reaction.emoji == '✅' or reaction.emoji == '⛔') and not member == botMember
    react = await bot.wait_for('reaction_add', check=check)
    if react[0].emoji == '✅':
        try:
            role = await server.create_role(name=name, colour=discord.Colour(int(hex,16)))
        except ValueError:
            role = await server.create_role(name=name, colour=discord.Colour(0x0))
        await ctx.author.add_roles(role)
        ws.cell(ws.max_row+1, 1, name)
        ws.cell(ws.max_row, 2, hex)
        wb.save('Role Colours.xlsx')
    for reaction in msg.reactions:
        await reaction.remove(botMember)
    


@bot.event
async def on_member_join(user):
    await generalChannel.send(user.mention + ''' has joined the server!
    Be sure to read the rules in ''' + ruleChannel.mention)


@bot.event
async def on_message(msg):
    await bot.process_commands(msg)
    if msg.content[0] == '`' and msg.channel is botChannel:
        roles = list(reversed([role.name for role in server.roles]))
        i = 1
        while ws.cell(i,1).value is not None:
            if ws.cell(i,1).value not in roles:
                roles += [ws.cell(i,1).value]
            i+=1
        i = 1
        found = 0
        for roleName in roles:
            if roleName.lower().startswith(msg.content[1:].lower()):
                if found: return
                role = roleName
                found = 1
        if found: await toggleRole(role, msg.author)


@bot.event
async def on_raw_reaction_add(payload):
    if payload.guild_id is None or payload.user_id == botMember.id: return
    emoji = payload.emoji.name
    channel = get(server.channels, id=payload.channel_id)
    member = payload.member
    if emoji == '🍆' or emoji == '🍑':
        msg = await channel.fetch_message(payload.message_id)
        if msg.author is botMember:
            await member.send('😉')

    if payload.message_id == pronounMsg:
        if emoji == '🟩':
            role = get(server.roles,name='He/Him')
        elif emoji == '🟥':
            role = get(server.roles,name='She/Her')
        elif emoji == '🟨':
            role = get(server.roles,name='They/Them')
        elif emoji == '🟪':
            role = get(server.roles,name='Ask-Pronouns')
        await member.add_roles(role)

    if payload.message_id == sexualityMsg:
        if emoji == 'Pride_Heart':
            role = get(server.roles,name='Gay')
        elif emoji == 'Lesbian_Heart':
            role = get(server.roles,name='Lesbian')
        elif emoji == 'Bi_Heart':
            role = get(server.roles,name='Bi')
        elif emoji == 'Pan_Heart':
            role = get(server.roles,name='Pan')
        elif emoji == 'Ace_Heart':
            role = get(server.roles,name='Ace')
        elif emoji == 'Aro_Heart':
            role = get(server.roles,name='Aro')
        elif emoji == '💝':
            role = get(server.roles,name='Straight')
        elif emoji == '💖':
            role = get(server.roles,name='Queer')
        elif emoji == '❓':
            role = get(server.roles,name='Questioning-Sexuality')
        await member.add_roles(role)

    if payload.message_id == genderMsg:
        if emoji == 'Trans_Heart':
            role = get(server.roles,name='Trans')
        elif emoji == '💙':
            role = get(server.roles,name='Male')
        elif emoji == '❤️':
            role = get(server.roles,name='Female')
        elif emoji == 'NonBinary_Heart':
            role = get(server.roles,name='Non-Binary')
        elif emoji == 'Agender_Heart':
            role = get(server.roles,name='Agender')
        elif emoji == 'Genderfluid_Heart':
            role = get(server.roles,name='Genderfluid')
        elif emoji == 'GenderQueer_Heart':
            role = get(server.roles,name='Genderqueer')
        elif emoji == '❓':
            role = get(server.roles,name='Questioning-Gender')
        await member.add_roles(role)

    if payload.message_id == enginMsg:
        if emoji == '1️⃣':
            role = 'Part-1'
        elif emoji == '2️⃣':
            role = 'Part-2'
        elif emoji == '3️⃣':
            role = 'Part-3'
        elif emoji == '4️⃣':
            role = 'Part-4'
        elif emoji == '🎓':
            role = 'Post-Grad'
        elif emoji == '🦠':
            role = 'Biomed'
        elif emoji == '🧪':
            role = 'Chemmat'
        elif emoji == '🏗️':
            role = 'Civil'
        elif emoji == '💾':
            role = 'CompSys'
        elif emoji == '⚡':
            role = 'Electrical'
        elif emoji == '🎲':
            role = 'EngSci'
        elif emoji == '⚙️':
            role = 'Mechanical'
        elif emoji == '🤖':
            role = 'Mechatronics'
        elif emoji == '🖥️':
            role = 'Software'
        await toggleRole(role, member)

        

@bot.event
async def on_raw_reaction_remove(payload):
    if payload.guild_id is None or payload.user_id == botMember.id: return
    emoji = payload.emoji.name
    channel = get(server.channels, id=payload.channel_id)
    member = get(server.members, id=payload.user_id)

    if payload.message_id == pronounMsg:
        if emoji == '🟩':
            role = get(server.roles,name='He/Him')
        elif emoji == '🟥':
            role = get(server.roles,name='She/Her')
        elif emoji == '🟨':
            role = get(server.roles,name='They/Them')
        elif emoji == '🟪':
            role = get(server.roles,name='Ask-Pronouns')
        await member.remove_roles(role)

    if payload.message_id == sexualityMsg:
        if emoji == 'Pride_Heart':
            role = get(server.roles,name='Gay')
        elif emoji == 'Lesbian_Heart':
            role = get(server.roles,name='Lesbian')
        elif emoji == 'Bi_Heart':
            role = get(server.roles,name='Bi')
        elif emoji == 'Pan_Heart':
            role = get(server.roles,name='Pan')
        elif emoji == 'Ace_Heart':
            role = get(server.roles,name='Ace')
        elif emoji == 'Aro_Heart':
            role = get(server.roles,name='Aro')
        elif emoji == '💝':
            role = get(server.roles,name='Straight')
        elif emoji == '💖':
            role = get(server.roles,name='Queer')
        elif emoji == '❓':
            role = get(server.roles,name='Questioning-Sexuality')
        await member.remove_roles(role)

    if payload.message_id == genderMsg:
        if emoji == 'Trans_Heart':
            role = get(server.roles,name='Trans')
        elif emoji == '💙':
            role = get(server.roles,name='Male')
        elif emoji == '❤️':
            role = get(server.roles,name='Female')
        elif emoji == 'NonBinary_Heart':
            role = get(server.roles,name='Non-Binary')
        elif emoji == 'Agender_Heart':
            role = get(server.roles,name='Agender')
        elif emoji == 'Genderfluid_Heart':
            role = get(server.roles,name='Genderfluid')
        elif emoji == 'GenderQueer_Heart':
            role = get(server.roles,name='Genderqueer')
        elif emoji == '❓':
            role = get(server.roles,name='Questioning-Gender')
        await member.remove_roles(role)

    if payload.message_id == enginMsg:
        if emoji == '1️⃣':
            role = 'Part-1'
        elif emoji == '2️⃣':
            role = 'Part-2'
        elif emoji == '3️⃣':
            role = 'Part-3'
        elif emoji == '4️⃣':
            role = 'Part-4'
        elif emoji == '🎓':
            role = 'Post-Grad'
        elif emoji == '🦠':
            role = 'Biomed'
        elif emoji == '🧪':
            role = 'Chemmat'
        elif emoji == '🏗️':
            role = 'Civil'
        elif emoji == '💾':
            role = 'CompSys'
        elif emoji == '⚡':
            role = 'Electrical'
        elif emoji == '🎲':
            role = 'EngSci'
        elif emoji == '⚙️':
            role = 'Mechanical'
        elif emoji == '🤖':
            role = 'Mechatronics'
        elif emoji == '🖥️':
            role = 'Software'
        await toggleRole(role, member)
    


async def toggleRole(role, member):
    roleObj = get(member.roles,name=role)
    if roleObj is not None:
        if roleObj.position > get(server.roles,name='Ask-Pronouns').position: return
        if len(roleObj.members) == 1 and roleObj.position < get(server.roles,name='Software').position:
            await roleObj.delete()
        else: await member.remove_roles(roleObj)
        return

    roleObj = get(server.roles, name=role)
    if roleObj is not None:
        if roleObj.position <= get(server.roles,name='Ask-Pronouns').position:
            await member.add_roles(roleObj)
        return
    i = 1
    while ws.cell(i,1).value is not None:
        if ws.cell(i,1).value == role:
            roleObj = await server.create_role(name=role, colour=discord.Colour(int(ws1.cell(i,2).value,16)))
            await member.add_roles(roleObj)
            return
        i += 1




bot.run(token)